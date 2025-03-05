import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copyfile
import os
import re
import time
import shutil
from openpyxl.worksheet.datavalidation import DataValidation
import numpy as np
from datetime import timedelta, datetime

# Read the worksheets into DataFrames
def worksheet_to_df(worksheet_name, columns_range):
    worksheet = spreadsheet.worksheet(worksheet_name)
    data = worksheet.get(columns_range)
    return pd.DataFrame(data[1:], columns=data[0])
    
    
# Setup Google Sheets API
scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('key.json', scope)
client = gspread.authorize(creds)
spreadsheet = client.open_by_url('https://docs.google.com/spreadsheets/d/1KN4XwXQlZ5jhyErxdpA_R6kNk2tKwwFeDnCDuHCr0fo/edit#gid=2033397596')
df_altas = worksheet_to_df('RAW_ALTA', 'A:AB')
df_ordenes = worksheet_to_df('RAW_IMSS', 'A:X')
df_gsheet_INSABI = worksheet_to_df('RAW_INSABI', 'A:AA')
df_zoho = worksheet_to_df('ZOHO', 'A:U')

print("\n*******************************\n  Homologando el formato de fechas.")

# Function to convert Excel serial dates and date-formatted strings to datetime
def excel_date_to_datetime(serial):
    if serial is None:
        return None
    date_pattern = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')  # Pattern to identify strings in date format

    # Attempt to exclude date-formatted strings and non-numeric values
    if isinstance(serial, str) and (date_pattern.match(serial) or not serial.replace(',', '').replace('.', '').isdigit()):
        try:
            # Convert string to date, specifying dayfirst=True to handle non-US date formats
            return pd.to_datetime(serial, errors='raise', dayfirst=True)
        except:
            return serial  # Return the original string if conversion fails

    try:
        # Remove commas for thousand separators and convert to float for processing
        serial = float(serial.replace(',', ''))
        excel_epoch_start = datetime(1899, 12, 30)
        if serial >= 1:
            return excel_epoch_start + timedelta(days=serial)
    except ValueError:
        return serial  # Return the original string if conversion fails

# Apply conversion across specified DataFrame columns and report issues
def convert_serial_dates_in_df(df, df_name, column_names):
    failed_conversions = []
    for column in column_names:
        if column in df.columns:
            # Apply conversion
            df[column] = df[column].apply(excel_date_to_datetime)
            # Attempt to normalize with pd.to_datetime, log failures
            original_non_na_count = df[column].notna().sum()
            df[column] = pd.to_datetime(df[column], errors='coerce')
            if df[column].notna().sum() < original_non_na_count:
                failed_conversions.append((df_name, column))

    # Report failed conversions
    if failed_conversions:
        for df_name, column in failed_conversions:
            print(f"Failed conversion in DataFrame '{df_name}', column '{column}'.")

# Example usage
# Ensure your dataframes, such as df_altas and df_gsheet_INSABI, are defined above this code.
date_columns = {
    'df_altas': ['fechaAltaTrunc', 'Fecha sistema', 'Fecha Pago KM'],
    'df_gsheet_INSABI': ['Fecha de ingreso', 'Fecha de pago', 'Fecha de entrega'],
}
for df_name, columns in date_columns.items():
    if df_name == 'df_altas':
        convert_serial_dates_in_df(df_altas, 'df_altas', columns)
    elif df_name == 'df_gsheet_INSABI':
        convert_serial_dates_in_df(df_gsheet_INSABI, 'df_gsheet_INSABI', columns)
print("Fechas homologadas \n*******************************\n")

# Agregando el Ciclo fiscal 

print("Agregando el ciclo fiscal") 
df_altas['fechaExpedicion'] = None
for index, row in df_altas.iterrows():
    order_match = df_ordenes[df_ordenes['orden'] == row['noOrden']]['fechaExpedicion'].values
    if order_match.size > 0:
        df_altas.at[index, 'fechaExpedicion'] = order_match[0]
df_altas['fechaExpedicion'] = df_altas['fechaExpedicion'].apply(lambda x: x[-4:] if pd.notnull(x) else x)
df_altas['fechaExpedicion'] = pd.to_numeric(df_altas['fechaExpedicion'], errors='coerce')
# Reemplaza los valores Problemáticos
factura_rewrite_dict = { "P-1347": "16/03/2023", "P-1348": "17/03/2023", "P-1350": "05/04/2023", "P-1396": "25/04/2023", "P-1390": "25/04/2023", "P-1389": "25/04/2023", "P-1391": "25/04/2023", "P-1395": "28/04/2023", "P-1405": "24/05/2023", "P-1417": "25/05/2023", "P-1422": "23/05/2023", "P-1413": "25/05/2023", "P-1424": "26/05/2023", "P-1432": "25/05/2023", "P-1431": "26/05/2023", "P-1441": "23/05/2023", "P-1528": "23/05/2023", "P-1454": "24/05/2023", "P-1446": "24/05/2023", "P-1525": "16/05/2023", "P-1451": "16/05/2023", "P-1444": "18/05/2023", "P-1439": "23/05/2023", "P-1465": "25/05/2023", "P-1464": "25/05/2023", "P-1450": "25/05/2023", "P-1443": "25/05/2023", "P-1463": "25/05/2023", "P-1484": "16/05/2023", "P-1497": "16/05/2023", "P-1488": "16/05/2023", "P-1487": "24/05/2023", "P-1478": "24/05/2023", "P-1496": "26/05/2023", "P-1499": "26/05/2023", "P-1494": "26/05/2023", "P-1492": "26/05/2023", "P-1491": "26/05/2023", "P-1493": "26/05/2023", "P-1508": "18/05/2023", "P-1502": "19/05/2023", "P-1506": "19/05/2023", "P-1509": "23/05/2023", "P-1513": "23/05/2023", "P-1517": "23/05/2023", "P-1516": "24/05/2023", "P-1530": "24/05/2023", "P-1534": "24/05/2023", "P-1532": "24/05/2023", "P-1589": "16/05/2023", "P-1537": "25/05/2023", "P-1609": "16/05/2023", "P-1616": "19/05/2023", "P-1624": "20/05/2023", "P-1612": "20/05/2023", "P-1634": "24/05/2023", "P-1628": "26/05/2023", "P-1615": "26/05/2023", "P-1632": "23/05/2023", "P-1620": "23/05/2023", "P-1629": "23/05/2023", "P-1640": "16/05/2023", "P-1613": "18/05/2023", "P-1608": "23/05/2023", "P-1645": "23/05/2023", "P-1646": "24/05/2023", "P-1677": "18/05/2023", "P-1694": "18/05/2023", "P-1644": "18/05/2023", "P-1669": "20/05/2023", "P-3139": "21/10/2023"}
for factura, fecha_sistema in factura_rewrite_dict.items():
    df_altas.loc[df_altas['Factura'] == factura, 'Fecha sistema'] = fecha_sistema
df_gsheet_INSABI['Ejercicio fiscal'] = df_gsheet_INSABI['FECHA EXPEDICIÓN DE LA ORDEN'].apply(lambda x: str(x)[-4:])


#Convierte a números los valores de importe
#print("Columnas:\n")
#print(df_altas.dtypes)
#print(df_ordenes.dtypes) 
#print(df_gsheet_INSABI.dtypes)
print("\n*******************************\n Conviertiendo tipos de datos")
    # Altas
df_altas['importe'] = pd.to_numeric(df_altas['importe'].str.replace('[$,]', '', regex=True), errors='coerce')
df_altas['Importe Pago KM'] = pd.to_numeric(df_altas['Importe Pago KM'].str.replace('[$,]', '', regex=True), errors='coerce')
df_altas['cantRecibida'] = pd.to_numeric(df_altas['cantRecibida'].str.replace('[$,]', '', regex=True), errors='coerce')
df_altas['fechaAltaTrunc'] = pd.to_datetime(df_altas['fechaAltaTrunc'], dayfirst=True, errors='coerce')
df_altas['Fecha sistema'] = pd.to_datetime(df_altas['Fecha sistema'], dayfirst=True, errors='coerce')
df_altas['Fecha Pago KM'] = pd.to_datetime(df_altas['Fecha Pago KM'], dayfirst=True, errors='coerce')

    # Órdenes
df_ordenes['cantidadSolicitada'] = pd.to_numeric(df_ordenes['cantidadSolicitada'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['precio'] = pd.to_numeric(df_ordenes['precio'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['importeSinIva'] = pd.to_numeric(df_ordenes['importeSinIva'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['% de cumplimiento'] = pd.to_numeric(df_ordenes['% de cumplimiento'], errors='coerce')
df_ordenes['Días a sancionar (piezas no entregadas)'] = pd.to_numeric(df_ordenes['Días a sancionar (piezas no entregadas)'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Sanción de piezas entregadas'] = pd.to_numeric(df_ordenes['Sanción de piezas entregadas'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Sanción de piezas no entregadas'] = pd.to_numeric(df_ordenes['Sanción de piezas no entregadas'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Sanciones totales'] = pd.to_numeric(df_ordenes['Sanciones totales'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Monto Entregado'] = pd.to_numeric(df_ordenes['Monto Entregado'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Parcialidad no entregada'] = pd.to_numeric(df_ordenes['Parcialidad no entregada'], errors='coerce')

    #INSABI
df_gsheet_INSABI['Importe'] = pd.to_numeric(df_gsheet_INSABI['Importe'].str.replace('[$,]', '', regex=True), errors='coerce')
df_gsheet_INSABI['Importe de pago'] = pd.to_numeric(df_gsheet_INSABI['Importe de pago'].str.replace('[$,]', '', regex=True), errors='coerce')
df_gsheet_INSABI['CANTIDAD SOLICITADA'] = pd.to_numeric(df_gsheet_INSABI['CANTIDAD SOLICITADA'], errors='coerce')
df_gsheet_INSABI['Ejercicio fiscal'] = pd.to_numeric(df_gsheet_INSABI['Ejercicio fiscal'], errors='coerce')
df_gsheet_INSABI['Fecha de ingreso'] = pd.to_datetime(df_gsheet_INSABI['Fecha de ingreso'], dayfirst=True, errors='coerce')
df_gsheet_INSABI['Fecha de pago'] = pd.to_datetime(df_gsheet_INSABI['Fecha de pago'], dayfirst=True, errors='coerce')
df_gsheet_INSABI['Fecha de entrega'] = pd.to_datetime(df_gsheet_INSABI['Fecha de entrega'], dayfirst=True, errors='coerce')

    #ZOHO
df_zoho['Importe'] = pd.to_numeric(df_zoho['Importe'], errors='coerce')
df_zoho['Ciclo fiscal'] = pd.to_numeric(df_zoho['Ciclo fiscal'], errors='coerce')
df_zoho['cantidadSolicitada'] = pd.to_numeric(df_zoho['cantidadSolicitada'], errors='coerce')
#print("Confirmando tipo de datos: ")
#print(df_altas.dtypes)
#print(df_ordenes.dtypes) 
#print(df_gsheet_INSABI.dtypes)
#print(df_zoho.dtypes)



print("\n*******************************\nDatos del GSHEET actualizados\n*******************************\n")

print("\n*******************************\n Tránsitos INSABI \n")
filter_32 = (
    (df_gsheet_INSABI['Ejercicio fiscal'] == 2024) &
    (df_gsheet_INSABI['Fuente de financiamiento'] == '32 Porciento') &
    (df_gsheet_INSABI['Fecha de entrega'].isna()) &
    (df_gsheet_INSABI['ESTATUS'] != 'Pendiente de remisión')
)
Sum_entransito32 = df_gsheet_INSABI.loc[filter_32, 'Importe'].sum()

# Filter for Sum_entransitoFonsabi
filter_Fonsabi = (
    (df_gsheet_INSABI['Ejercicio fiscal'] == 2024) &
    (df_gsheet_INSABI['Fuente de financiamiento'] == 'FONSABI') &
    (df_gsheet_INSABI['Fecha de entrega'].isna()) &
    (df_gsheet_INSABI['ESTATUS'] != 'Pendiente de remisión')
)
Sum_entransitoFonsabi = df_gsheet_INSABI.loc[filter_Fonsabi, 'Importe'].sum()

# Print results
print(f"Sum_entransito32: {Sum_entransito32}")
print(f"Sum_entransitoFonsabi: {Sum_entransitoFonsabi}")

print("\n*******************************\n")

# Assuming df_altas, df_ordenes, df_gsheet_INSABI are already defined and loaded with data

# Initialize an empty DataFrame for df_summary
df_altas['importe_numeric'] = df_altas['importe'].replace('[\$,]', '', regex=True).astype(float)
df_altas['Importe Pago KM_numeric'] = pd.to_numeric(df_altas['Importe Pago KM'].replace('[\$,]', '', regex=True), errors='coerce')
df_altas['Importe Pago KM_numeric'].fillna(0, inplace=True)


df_summary = pd.DataFrame(index=["Sum Altas", "Sum Contrarecibos", "Pagado"], columns=['U230178', 'U220749', 'U210607'])

# Sumar altas
for noContrato in df_summary.columns:
    df_summary.at["Sum Altas", noContrato] = df_altas[df_altas['noContrato'] == noContrato]['importe_numeric'].sum()

# Sumar contrarecibos
pisp_conditions = ["Pagado", "Aprobado", "En proceso"]
for noContrato in df_summary.columns:
    sum_value = df_altas[(df_altas['noContrato'] == noContrato) & (df_altas['PISP'].isin(pisp_conditions))]['importe_numeric'].sum()
    df_summary.at["Sum Contrarecibos", noContrato] = sum_value

# Pagado
for noContrato in df_summary.columns:
    sum_value = df_altas.loc[
        (df_altas['noContrato'] == noContrato) &
        (~df_altas['Fecha Pago KM'].isna()),  # Directly check for non-NaN (i.e., valid datetime objects)
        'importe_numeric'
    ].sum()
    df_summary.at["Pagado", noContrato] = sum_value

# Since 'Fecha Pago KM' is already a datetime, there's no need for date pattern filtering here either
filtered_importe_numeric = df_altas.loc[~df_altas['Fecha Pago KM'].isna(), 'importe_numeric'].sum()
# Sum of the whole 'Importe Pago KM_numeric' column
total_importe_pago_km_numeric = df_altas['Importe Pago KM_numeric'].sum()
# Calculating 'Sanciones' using the filtered sum
sanciones = filtered_importe_numeric - total_importe_pago_km_numeric

# Print the df_summary and other information
print(df_summary)
print(f"Facturas pagadas {filtered_importe_numeric}")
print(f"Importe pagado {total_importe_pago_km_numeric}")
print(f"Sanciones: {sanciones}")

# Vamos a crear el excel 
# Define the source file path and the destination file path with the current date and hour
source_file_path = './templatereportecobros.xlsx'
current_date_hour = datetime.now().strftime("%m-%d-%Y %Hh")
new_file_name = f"Reporte cobros {current_date_hour}.xlsx"
destination_file_path = f"./{new_file_name}"

# Copy the file
shutil.copy(source_file_path, destination_file_path)

# Load the workbook and select the "Resumen" sheet
wb = load_workbook(destination_file_path)
ws = wb["Resumen"]

# Fecha del reporte
current_date = datetime.now().strftime("%d-%b-%Y")  # Format: dd-mmm-yyyy
ws['C1'] = current_date
ws['D7'] = sanciones


# Mercancía en tránsito
#Convierte a numérico
df_ordenes['importeSinIva'] = pd.to_numeric(df_ordenes['importeSinIva'], errors='coerce')

# This step depends on the actual content of the column.
df_ordenes['¿Orden cancelada?'] = df_ordenes['¿Orden cancelada?'].map({'FALSE': False, 'True': True})

# Assuming 'Monto Entregado' also needs to be numeric for your condition,
df_ordenes['Monto Entregado'] = pd.to_numeric(df_ordenes['Monto Entregado'], errors='coerce')
df_ordenesSummary = pd.DataFrame(index=["Mercancía en tránsito"], columns=['U230178', 'U220749', 'U210607'])
# Sum "importeSinIva" column
for contrato in df_ordenesSummary.columns:
    df_ordenesSummary.at["Mercancía en tránsito", contrato] = df_ordenes[
        (df_ordenes['contrato'] == contrato) &  # Use the loop variable directly
        (df_ordenes['¿Orden cancelada?'] == False) &
        (df_ordenes['estatus'] == 'Confirmada') &
        (df_ordenes['Monto Entregado'] == 0)
    ]['importeSinIva'].sum()
wb.save(destination_file_path)
# Load the workbook and select the relevant sheet
wb = load_workbook(destination_file_path)
ws = wb["Resumen"]

# Assuming df_ordenesSummary is already defined
# Prepare data for writing
data_to_write = df_ordenesSummary.values  # Use the entire DataFrame

# Mercancía en tránsito
start_row = 11  # Starting at row 11
start_column = 3  # Starting at column C, corresponding to 3

for row_index, row in enumerate(data_to_write, start=start_row):
    for col_index, value in enumerate(row, start=start_column):
        ws.cell(row=row_index, column=col_index, value=value)



# Agrega el desglose IMSS
sheet_name = "Desglose IMSS"
if sheet_name in wb.sheetnames:
    ws_desglose = wb[sheet_name]
    # Clear the content but leave the structure
    for row in ws_desglose.iter_rows(min_row=2, max_col=ws_desglose.max_column, max_row=ws_desglose.max_row):
        for cell in row:
            cell.value = None
else:
    ws_desglose = wb.create_sheet(sheet_name)



# Define column headers and map them to df_altas columns
headers = {
    "Fecha de entrega de la orden": "fechaAltaTrunc",
    "Importe": "importe",
    "Factura": "Factura",
    "Fecha de contrarecibo": "Fecha sistema",
    "Días Entrega-Contrarecibo": None,  # Formula to be added later
    "Fecha de pago": "Fecha Pago KM",
    "Días Contrarecibo-Pago": None,  # Formula to be added later
    "Ciclo fiscal": "fechaExpedicion",
    "Contrato": "noContrato"
}

# Add headers to the second row
ws_desglose.append([None] * (len(headers) + 1))  # Leave the first row entirely blank
ws_desglose.append(list(headers.keys()))

# Populate the sheet with data from df_altas
for r_index, row in enumerate(dataframe_to_rows(df_altas, index=False, header=False), start=3):
    # Prepare data row based on headers mapping
    data_row = [row[df_altas.columns.get_loc(headers[h])] if headers[h] else None for h in headers]
    ws_desglose.append(data_row)

# Adding formulas where applicable
for row in range(3, ws_desglose.max_row + 1):
    # Column E formula (Días Entrega-Contrarecibo)
    ws_desglose[f'E{row}'] = f'=IF(D{row}="",TODAY()-A{row},D{row}-A{row})'
    # Column G formula (Días Contrarecibo-Pago)
    ws_desglose[f'G{row}'] = f'=IF(AND(F{row}="",D{row}=""),"",IF(F{row}="",TODAY()-D{row},F{row}-D{row}))'

ws_desglose['A1'] = f'=SUBTOTAL(9, B3:B{ws_desglose.max_row})'


# Print the summary DataFrame
print(df_ordenesSummary)
# Save the workbook
wb.save(destination_file_path)

########### Aquí empieza el desglose INSABI ##############

# Sanciones antes de ligar el oficio 
# Sanciones FONSABI
# Sanciones 32%
#porciento32_filtered = df_gsheet_INSABI[(df_gsheet_INSABI['Fuente de financiamiento'] == "32 Porciento") & (df_gsheet_INSABI['Fecha de pago'].notna())]
#fonsabi_filtered = df_gsheet_INSABI[(df_gsheet_INSABI['Fuente de financiamiento'] == "FONSABI") & (df_gsheet_INSABI['Fecha de pago'].notna())]
df_gsheet_INSABI['sanción'] = pd.to_numeric(df_gsheet_INSABI['sanción'], errors='coerce').fillna(0)


sanciones32porciento = df_gsheet_INSABI[df_gsheet_INSABI['Fuente de financiamiento'] == "32 Porciento"]['sanción'].sum()
# Calculate the sum of 'Sanción' for 'FONSABI'
sancionesFONSABI = df_gsheet_INSABI[df_gsheet_INSABI['Fuente de financiamiento'] == "FONSABI"]['sanción'].sum()
# Print results
print("\n*******************************\n Sanciones INSABI \n")
print(f"Sanciones FONSABI: {sancionesFONSABI}")
print(f"Sanciones 32 Porciento: {sanciones32porciento}")
print("\n*******************************\n")
df_gsheet_INSABI = df_gsheet_INSABI[
    df_gsheet_INSABI['Fecha de entrega'].notna()
]

wb = load_workbook(destination_file_path)

sheet_name = "Desglose INSABI"
if sheet_name in wb.sheetnames:
    ws_desglose = wb[sheet_name]
    for row in ws_desglose.iter_rows(min_row=2, max_col=ws_desglose.max_column, max_row=ws_desglose.max_row):
        for cell in row:
            cell.value = None
else:
    ws_desglose = wb.create_sheet(sheet_name)

# Adjust headers if necessary to match your DataFrame columns
headers = {
    "Fecha de entrega de la orden": "Fecha de entrega",
    "Importe": "Importe",
    "Factura": "Factura",
    "Fecha de contrarecibo": "Fecha de ingreso",
    "Días Entrega-Contrarecibo": None,  # Formula to be added later
    "Fecha de pago": "Fecha de pago",
    "Días Contrarecibo-Pago": None,  # Formula to be added later
    "Ciclo fiscal": "Ejercicio fiscal",
    "Contrato": "Fuente de financiamiento"
}

ws_desglose.append([None] * (len(headers) + 1))  # Adjust if you want to leave the first row entirely blank
ws_desglose.append(list(headers.keys()))

for row in dataframe_to_rows(df_gsheet_INSABI, index=False, header=False):
    data_row = []
    for h in headers:
        if headers[h]:
            try:
                data_row.append(row[df_gsheet_INSABI.columns.get_loc(headers[h])])
            except KeyError:
                data_row.append(None)  # Append None if column not found
        else:
            data_row.append(None)  # Append None for columns with formulas to be added later
    ws_desglose.append(data_row)

for row in range(3, ws_desglose.max_row + 1):
    ws_desglose[f'E{row}'] = f'=IF(D{row}="",TODAY()-A{row},D{row}-A{row})'
    ws_desglose[f'G{row}'] = f'=IF(AND(F{row}="",D{row}=""),"",IF(F{row}="",TODAY()-D{row},F{row}-D{row}))'

# Assuming 'B1' is intended to be a summary or total calculation; adjust as needed
ws_desglose['A1'] = f'=SUBTOTAL(9, B3:B{ws_desglose.max_row})'
wb.save(destination_file_path)

#wb = load_workbook(destination_file_path)
ws = wb["Resumen"]

ws['D16'] = sanciones32porciento
ws['E16'] = sancionesFONSABI
ws['D21'] = Sum_entransito32
ws['E21'] = Sum_entransitoFonsabi


wb.save(destination_file_path)

########### Aquí empieza el desglose CCINSHAE ##############
"""
# Initial state of specific 'Factura' values
print("Initial state:")
print(df_zoho[df_zoho['Factura'].isin(['P-1474', 'P-1655', 'P-4258'])][['Factura', 'Estatus Factura']])
# Replace "#N/A" with NaN across all columns
df_zoho.replace("#N/A", np.nan, inplace=True)
# Ensure 'Factura' and 'Estatus Factura' are treated as strings
df_zoho['Factura'] = df_zoho['Factura'].astype(str)
df_zoho['Estatus Factura'] = df_zoho['Estatus Factura'].astype(str).str.strip()
# Replace "Cancelado" or blank with NaN in "Estatus Factura"
# Apply the operation without using inplace=True in a chained assignment
df_zoho['Estatus Factura'] = df_zoho['Estatus Factura'].replace(["Cancelado", ""], np.nan)
# Replace empty strings with NaN to catch all empty values in 'Factura'
df_zoho['Factura'] = df_zoho['Factura'].replace('', np.nan)
# Drop rows where 'Factura' is empty or NaN
df_zoho.dropna(subset=['Factura'], inplace=True)
# Drop rows where "Estatus Factura" is NaN (including those set from "Cancelado" or blank)
df_zoho.dropna(subset=['Estatus Factura'], inplace=True)

# Function to check if all values are the same in a column for a given group
def check_values(series, factura_value):
    unique_values = series.unique()
    if len(unique_values) > 1:
        print(f"Warning: {len(unique_values)} unique values for '{factura_value}' in column '{series.name}'.")
        print(f"Unique values include: {unique_values[:5]}")  # Print first 5 unique values as an example
        # Print 'Factura' value for context
        print(f"Corresponding 'Factura' value: {factura_value}")

# Example to manually check for discrepancies before aggregation
for factura, group in df_zoho.groupby('Factura'):
    for column in ['Fecha de entrega de la orden', 'Fecha de pago', 'Ciclo fiscal', 'Contrato', 'Institución']:
        if group[column].nunique() > 1:
            print(f"Warning: Multiple unique values for '{factura}' in column '{column}'.")
            print(f"Unique values include: {group[column].unique()[:5]}")  # Show a sample of unique values

# Step 3: Group by 'Factura' and process
aggregation = {
    'Importe': 'sum',
    'Fecha de entrega de la orden': lambda x: check_values(x, x.name) or x.iloc[0],
    'Fecha de pago': lambda x: check_values(x, x.name) or x.iloc[0],
    'Ciclo fiscal': lambda x: check_values(x, x.name) or x.iloc[0],
    'Contrato': lambda x: check_values(x, x.name) or x.iloc[0],
    'lugarEntrega': lambda x: check_values(x, x.name) or x.iloc[0],
}
# Group by 'Factura' and apply aggregation
df_zoho = df_zoho.groupby('Factura', as_index=False).agg(aggregation)

wb = load_workbook(destination_file_path)
sheet_name = "Desglose CCINSHAE"

# Ensure the sheet exists or create it
if sheet_name in wb.sheetnames:
    ws_desgloseccinshae = wb[sheet_name]
    # Clear the content but leave the structure
    for row in ws_desgloseccinshae.iter_rows(min_row=2, max_col=ws_desgloseccinshae.max_column, max_row=ws_desgloseccinshae.max_row):
        for cell in row:
            cell.value = None
else:
    ws_desgloseccinshae = wb.create_sheet(sheet_name)

# Define column headers and map them to df_zoho columns
headers = {
    "Fecha de entrega de la orden": "Fecha de entrega de la orden",
    "Importe": "Importe",
    "Factura": "Factura",
    "Fecha de contrarecibo": 'Fecha de Ingreso',  # Leave blank
    "Días Entrega-Contrarecibo": None,  # Leave blank
    "Fecha de pago": "Fecha de pago",
    "Días Contrarecibo-Pago": None,  # Leave blank
    "Ciclo fiscal": "Ciclo fiscal",
    "Contrato": "Contrato",
    "Institución": "lugarEntrega"
}

# Add headers to the sheet
ws_desgloseccinshae.append([None] * (len(headers) + 1))  # Optional: Adjust if you need to leave the first column blank
ws_desgloseccinshae.append(list(headers.keys()))


# Populate the sheet with data from df_zoho
for r_index, row in enumerate(dataframe_to_rows(df_zoho, index=False, header=False), start=3):
    data_row = [row[df_zoho.columns.get_loc(h)] if h in df_zoho.columns else None for h in headers.values()]
    ws_desgloseccinshae.append(data_row)

# Adding formulas where applicable
for row in range(3, ws_desgloseccinshae.max_row + 1):
    ws_desgloseccinshae[f'E{row}'] = f'=IF(D{row}="",TODAY()-A{row},D{row}-A{row})'
    ws_desgloseccinshae[f'G{row}'] = f'=IF(AND(F{row}="",D{row}=""),"",IF(F{row}="",TODAY()-D{row},F{row}-D{row}))'

# Update the formula in A1 if needed
ws_desgloseccinshae['A1'] = f'=SUBTOTAL(9, B3:B{ws_desgloseccinshae.max_row})'

# Save the workbook
wb.save(destination_file_path)
"""

print("\n*******************************\nReportes finalizados\n*******************************\n")