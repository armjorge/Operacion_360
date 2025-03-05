import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
from datetime import datetime
import os
import shutil
from openpyxl import load_workbook



# Setup Google Sheets API
scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('key.json', scope)
client = gspread.authorize(creds)
spreadsheet = client.open_by_url('https://docs.google.com/spreadsheets/d/1KN4XwXQlZ5jhyErxdpA_R6kNk2tKwwFeDnCDuHCr0fo/edit#gid=2033397596')

# Read the worksheets into DataFrames
def worksheet_to_df(worksheet_name, columns_range):
    worksheet = spreadsheet.worksheet(worksheet_name)
    data = worksheet.get(columns_range)
    return pd.DataFrame(data[1:], columns=data[0])

df_altas = worksheet_to_df('RAW_ALTA', 'A:AB')
df_ordenes = worksheet_to_df('RAW_IMSS', 'A:X')
df_gsheet_INSABI = worksheet_to_df('RAW_INSABI', 'A:Y')
df_PAQ_ZOHO = worksheet_to_df('PAQ_ZOHO', 'A:T')
# Esto es de IMSS
# Agregando el Ciclo fiscal 
print("Agregando el ciclo fiscal") 
df_altas['fechaExpedicion'] = None
for index, row in df_altas.iterrows():
    order_match = df_ordenes[df_ordenes['orden'] == row['noOrden']]['fechaExpedicion'].values
    if order_match.size > 0:
        df_altas.at[index, 'fechaExpedicion'] = order_match[0]
df_altas['fechaExpedicion'] = df_altas['fechaExpedicion'].apply(lambda x: x[-4:] if pd.notnull(x) else x)
# Reemplaza los valores Problemáticos
factura_rewrite_dict = { "P-1347": "16/03/2023", "P-1348": "17/03/2023", "P-1350": "05/04/2023", "P-1396": "25/04/2023", "P-1390": "25/04/2023", "P-1389": "25/04/2023", "P-1391": "25/04/2023", "P-1395": "28/04/2023", "P-1405": "24/05/2023", "P-1417": "25/05/2023", "P-1422": "23/05/2023", "P-1413": "25/05/2023", "P-1424": "26/05/2023", "P-1432": "25/05/2023", "P-1431": "26/05/2023", "P-1441": "23/05/2023", "P-1528": "23/05/2023", "P-1454": "24/05/2023", "P-1446": "24/05/2023", "P-1525": "16/05/2023", "P-1451": "16/05/2023", "P-1444": "18/05/2023", "P-1439": "23/05/2023", "P-1465": "25/05/2023", "P-1464": "25/05/2023", "P-1450": "25/05/2023", "P-1443": "25/05/2023", "P-1463": "25/05/2023", "P-1484": "16/05/2023", "P-1497": "16/05/2023", "P-1488": "16/05/2023", "P-1487": "24/05/2023", "P-1478": "24/05/2023", "P-1496": "26/05/2023", "P-1499": "26/05/2023", "P-1494": "26/05/2023", "P-1492": "26/05/2023", "P-1491": "26/05/2023", "P-1493": "26/05/2023", "P-1508": "18/05/2023", "P-1502": "19/05/2023", "P-1506": "19/05/2023", "P-1509": "23/05/2023", "P-1513": "23/05/2023", "P-1517": "23/05/2023", "P-1516": "24/05/2023", "P-1530": "24/05/2023", "P-1534": "24/05/2023", "P-1532": "24/05/2023", "P-1589": "16/05/2023", "P-1537": "25/05/2023", "P-1609": "16/05/2023", "P-1616": "19/05/2023", "P-1624": "20/05/2023", "P-1612": "20/05/2023", "P-1634": "24/05/2023", "P-1628": "26/05/2023", "P-1615": "26/05/2023", "P-1632": "23/05/2023", "P-1620": "23/05/2023", "P-1629": "23/05/2023", "P-1640": "16/05/2023", "P-1613": "18/05/2023", "P-1608": "23/05/2023", "P-1645": "23/05/2023", "P-1646": "24/05/2023", "P-1677": "18/05/2023", "P-1694": "18/05/2023", "P-1644": "18/05/2023", "P-1669": "20/05/2023", "P-3139": "21/10/2023"}
for factura, fecha_sistema in factura_rewrite_dict.items():
    df_altas.loc[df_altas['Factura'] == factura, 'Fecha sistema'] = fecha_sistema

df_gsheet_INSABI['Ejercicio fiscal'] = df_gsheet_INSABI['FECHA EXPEDICIÓN DE LA ORDEN'].apply(lambda x: str(x)[-4:])

#Convierte a números los valores de importe
print("Columnas:\n")
print(df_altas.dtypes)
print(df_ordenes.dtypes) 
print(df_gsheet_INSABI.dtypes)
print("Conviertiendo tipos de datos")
    # Altas
df_altas['importe'] = pd.to_numeric(df_altas['importe'].str.replace('[$,]', '', regex=True), errors='coerce')
df_altas['Importe Pago KM'] = pd.to_numeric(df_altas['Importe Pago KM'].str.replace('[$,]', '', regex=True), errors='coerce')
df_altas['cantRecibida'] = pd.to_numeric(df_altas['cantRecibida'].str.replace('[$,]', '', regex=True), errors='coerce')
df_altas['fechaExpedicion'] = pd.to_numeric(df_altas['fechaExpedicion'], errors='coerce')

    # Órdenes
df_ordenes['cantidadSolicitada'] = pd.to_numeric(df_ordenes['cantidadSolicitada'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['precio'] = pd.to_numeric(df_ordenes['precio'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['importeSinIva'] = pd.to_numeric(df_ordenes['importeSinIva'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['% de cumplimiento'] = pd.to_numeric(df_ordenes['% de cumplimiento'], errors='coerce')
df_ordenes['Días a sancionar (piezas no entregadas)'] = pd.to_numeric(df_ordenes['Días a sancionar (piezas no entregadas)'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Sanción de piezas entregadas'] = pd.to_numeric(df_ordenes['Sanción de piezas entregadas'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Sanción de piezas no entregadas'] = pd.to_numeric(df_ordenes['Sanción de piezas no entregadas'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Sanciones totales'] = pd.to_numeric(df_ordenes['Sanciones totales'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Monto Entregado'] = pd.to_numeric(df_ordenes['Monto Entregado'].str.replace('[$,]', '', regex=True), errors='coerce')
df_ordenes['Parcialidad no entregada'] = pd.to_numeric(df_ordenes['Parcialidad no entregada'], errors='coerce')

    #INSABI
df_gsheet_INSABI['Importe'] = pd.to_numeric(df_gsheet_INSABI['Importe'].str.replace('[$,]', '', regex=True), errors='coerce')
df_gsheet_INSABI['Importe de pago'] = pd.to_numeric(df_gsheet_INSABI['Importe de pago'].str.replace('[$,]', '', regex=True), errors='coerce')
df_gsheet_INSABI['CANTIDAD SOLICITADA'] = pd.to_numeric(df_gsheet_INSABI['CANTIDAD SOLICITADA'], errors='coerce')
df_gsheet_INSABI['Ejercicio fiscal'] = pd.to_numeric(df_gsheet_INSABI['Ejercicio fiscal'], errors='coerce')

print("Confirmando tipo de datos: ")
print(df_altas.dtypes)
print(df_ordenes.dtypes) 
print(df_gsheet_INSABI.dtypes)

print("\n*******************************\nDatos del GSHEET actualizados\n*******************************\n")

####################### A partir de aquí puedes usar los datos como te plazca############

headers = ['Nombre cliente', 'Unidad operativa', 'Factura', 'Ejercicio fiscal', 'Fecha factura', 'Importe', 'Alta', 'base', 'Contrato', 'Fecha de pago CR', 'UUID']
df_reporteTYCSA = pd.DataFrame(columns=headers)

# Populate df_reporteTYCSA with data from df_altas
df_altas_mapped = pd.DataFrame({
    'Nombre cliente': "IMSS ESEOTRES",
    'Unidad operativa': df_altas['descUnidad'],
    'Factura': df_altas['Factura'],
    'Ejercicio fiscal': df_altas['fechaExpedicion'],
    'Fecha factura': df_altas['Fecha factura'],
    'Importe': df_altas['importe'],
    'Alta': df_altas['noAlta'],
    'base': df_altas['noOrden'],
    'Contrato': df_altas['noContrato'],
    'Fecha de pago CR': df_altas['Fecha Pago KM'],
    'UUID': df_altas['Folio']
})

# Append df_altas_mapped to df_reporteTYCSA
df_reporteTYCSA = pd.concat([df_reporteTYCSA, df_altas_mapped], ignore_index=True)

# Map and append data from df_gsheet_INSABI
#df_gsheet_INSABI['Ejercicio fiscal'] = df_gsheet_INSABI['FECHA EXPEDICIÓN DE LA ORDEN'].apply(lambda x: str(x)[-4:])

df_gsheet_INSABI_mapped = pd.DataFrame({
    'Nombre cliente': "IMSS-BIENESTAR ESEOTRES",
    'Unidad operativa': "IMSS-BIENESTAR",
    'Factura': df_gsheet_INSABI['Factura'],
    'Ejercicio fiscal': df_gsheet_INSABI['Ejercicio fiscal'],
    'Fecha factura': df_gsheet_INSABI['Fecha factura'],
    'Importe': df_gsheet_INSABI['Importe'],
    'Alta': '',
    'base': df_gsheet_INSABI['NÚMERO DE ORDEN DE SUMINISTRO'],
    'Contrato': df_gsheet_INSABI['Contrato'],
    'Fecha de pago CR': df_gsheet_INSABI['Fecha de pago'],
    'UUID': df_gsheet_INSABI['UUID']
})

# Append df_gsheet_INSABI_mapped to df_reporteTYCSA
df_reporteTYCSA = pd.concat([df_reporteTYCSA, df_gsheet_INSABI_mapped], ignore_index=True)

df_reporteTYCSA = df_reporteTYCSA[~df_reporteTYCSA['Factura'].isin(['#N/A', '']) & df_reporteTYCSA['Factura'].notnull()]

# Agregar los CCINSHAES
df_PAQ_ZOHO_mapped = pd.DataFrame({
    'Nombre cliente': df_PAQ_ZOHO['Institución Homologada'],
    'Unidad operativa': df_PAQ_ZOHO['lugarEntrega'],
    'Factura': df_PAQ_ZOHO['Factura'],
    'Ejercicio fiscal': df_PAQ_ZOHO['Ciclo fiscal'],
    'Fecha factura': df_PAQ_ZOHO['Fecha'],
    'Importe': df_PAQ_ZOHO['Total'],
    'Alta': df_PAQ_ZOHO['Referencia'],
    'base': df_PAQ_ZOHO['Orden'],
    'Contrato': df_PAQ_ZOHO['Contrato'],
    'Fecha de pago CR': df_PAQ_ZOHO['Fecha Pago'],
    'UUID': df_PAQ_ZOHO['UUID']
})

df_reporteTYCSA = pd.concat([df_reporteTYCSA, df_PAQ_ZOHO_mapped], ignore_index=True)

############# REPORTE A EXCEL ##############
print("\n*******************************\nGenerando Excel\n*******************************\n")

file_path = './M02_cartera.xlsx'
sheet_name = 'DETALLE'
current_date = datetime.now().strftime("%m-%d-%Y")
new_file_path = file_path.replace('.xlsx', f'_{current_date}.xlsx')

# Copying the original Excel file to a new one with the current date
shutil.copy(file_path, new_file_path)

# Assuming df_reporteTYCSA is your DataFrame and is already defined
# Make sure to load or define your DataFrame here

# Column mapping from df_reporteTYCSA to Excel
column_mapping = {
    'Nombre cliente': 'Nombre cliente',
    'Unidad operativa': 'Unidad operativa',
    'Factura': 'Factura',
    'Ejercicio fiscal': 'Ejercicio fiscal',
    'Fecha factura': 'Fecha factura',
    'Importe': 'Importe',
    'Alta': 'Alta',
    'base': 'base',
    'Contrato': 'Contrato',
    'Fecha de pago CR': 'Fecha de pago CR',
    'UUID': 'UUID',
}

# Open the workbook and target sheet
wb = load_workbook(new_file_path)
ws = wb[sheet_name]

# Finding column indices in the Excel sheet based on header row
header_row = ws[1]
excel_column_indices = {cell.value: cell.column for cell in header_row}

# Iterate over each mapping and update Excel accordingly
for df_col, excel_col in column_mapping.items():
    if df_col in df_reporteTYCSA.columns and excel_col in excel_column_indices:
        # Determine the Excel column index
        col_index = excel_column_indices[excel_col]
        for row_index, value in enumerate(df_reporteTYCSA[df_col], start=2):  # Assuming data starts at row 2
            ws.cell(row=row_index, column=col_index, value=value)

# Save and close the workbook
wb.save(new_file_path)
wb.close()