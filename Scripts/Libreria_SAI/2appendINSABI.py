import pandas as pd
import glob
import os
from openpyxl import load_workbook

# Delete the file 'ordenesSuministro.xlsx' if it exists
output_file_name = 'ordenesSuministro.xlsx'
if os.path.exists(output_file_name):
    os.remove(output_file_name)
    print(f"{output_file_name} has been deleted.")
else:
    print(f"No such file: {output_file_name}")

# Get a list of all the excel files in the current directory with the prefix 'ordenesSuministro'
excel_files = glob.glob('ordenesSuministro*.xlsx')

# Read them into pandas dataframes and concatenate them into one dataframe
df = pd.concat((pd.read_excel(file) for file in excel_files))
file_path = r"C:\Users\armjorge\Dropbox\3. Armando Cuaxospa\Reportes GPT\Dataframes\Camunda\INSABI_ordenes-contratos-sellos-remisiones.xlsx"
sheet_name = 'Órdenes'

try:
    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)
    book.save(file_path)
except FileNotFoundError:
    print("File not found. A new file will be created.")
    
# Check for duplicates
if df.duplicated().any():
    print("\n******************************* \nValores fuente duplicados, descargar de nuevo la información fuente.\n*******************************\n")
else:
    # Save the combined dataframe to a new excel file
    df.to_excel('ordenesSuministro.xlsx', index=False)
    print("\n******************************* \nArchivos fusionados exitosamente\n**********")
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"\n******************************* \nArchivos agregado a '{sheet_name}' in '{file_path}' exitosamente.\n******************************* \n")

