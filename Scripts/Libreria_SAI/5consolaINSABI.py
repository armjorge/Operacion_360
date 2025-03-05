import openpyxl

def copy_data(source_sheet, target_sheet):
    for i, row in enumerate(source_sheet.iter_rows(), start=1):
        for j, cell in enumerate(row, start=1):
            target_sheet.cell(row=i, column=j, value=cell.value)

# Load the workbooks
ordenes_insabi_wb = openpyxl.load_workbook("ordenesSuministro.xlsx")
remisiones_insabi_wb = openpyxl.load_workbook("remisiones_INSABI.xlsx")

# Get the sheets
ordenes_sheet = ordenes_insabi_wb.active
remisiones_sheet = remisiones_insabi_wb.active

# Create a new workbook
consola_insabi_wb = openpyxl.Workbook()

# Create and name the sheets
ordenes_target_sheet = consola_insabi_wb.active
ordenes_target_sheet.title = "Órdenes"

remisiones_target_sheet = consola_insabi_wb.create_sheet("Remisiones")

# Copy data from ordenesSuministro.xlsx to ConsolaINSABI.xlsx's "Órdenes" sheet
copy_data(ordenes_sheet, ordenes_target_sheet)

# Copy data from remisiones_INSABI.xlsx to ConsolaINSABI.xlsx's "Remisiones" sheet
copy_data(remisiones_sheet, remisiones_target_sheet)

# Save the consola_insabi workbook
consola_insabi_wb.save("ConsolaINSABI.xlsx")
