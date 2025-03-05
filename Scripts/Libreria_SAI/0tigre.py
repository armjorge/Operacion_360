import openpyxl
import glob
import os

# Define a function to get the most recently modified file with a certain prefix
def get_latest_file(prefix):
    files = glob.glob(f"{prefix}*.xlsx")
    if not files:
        raise FileNotFoundError(f"No file with prefix {prefix} found")
    latest_file = max(files, key=os.path.getctime)
    return latest_file

# Load the workbooks
tigre_wb = openpyxl.load_workbook("tigre_2024.xlsx")

# Use the function to find and load the latest files with specified prefixes
ordenes_wb = openpyxl.load_workbook(get_latest_file('ordenes_export_'))
altas_wb = openpyxl.load_workbook(get_latest_file('altas_export_'))

# Get the sheets
tigre_ordenes = tigre_wb["Órdenes"]
tigre_altas = tigre_wb["Altas"]
#tigre_excel_de_trabajo = tigre_wb["Excel de trabajo"]
ordenes_sheet = ordenes_wb.active
altas_sheet = altas_wb.active

# Clear the "Órdenes" and "Altas" sheets
for sheet in [tigre_ordenes, tigre_altas]:
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

# Copy from ordenes.xlsx to tigre.xlsx's "Órdenes" sheet
for i, row in enumerate(ordenes_sheet.iter_rows(), start=1):
    for j, cell in enumerate(row, start=1):
        tigre_ordenes.cell(row=i, column=j, value=cell.value)

# Convert the values in column C of the "Órdenes" sheet from strings to integers
for cell in tigre_ordenes['C'][1:]:  # Skip the header
    if cell.value is not None:
        try:
            cell.value = int(cell.value)
        except ValueError:
            continue  # skips to the next iteration if the value cannot be converted to an integer

# Copy from altas.xlsx to tigre.xlsx's "Altas" sheet
for i, row in enumerate(altas_sheet.iter_rows(), start=1):
    for j, cell in enumerate(row, start=1):
        tigre_altas.cell(row=i, column=j, value=cell.value)

# Copy from altas.xlsx to "Excel de trabajo" sheet's column A
#for i, cell in enumerate(tigre_altas['A'], start=1):
 #   if cell.value is not None:
  #      tigre_excel_de_trabajo.cell(row=i, column=1, value=cell.value)
"""
# Set the formulas for the other columns in "Excel de trabajo"
last_row = tigre_excel_de_trabajo.max_row
#formulas = [
   '=INDEX(Altas!D:D,MATCH(A2,Altas!A:A,0))',
    '=INDEX(Órdenes!A:A,MATCH(B2,Órdenes!C:C,0))',
    '=INDEX(Altas!F:F,MATCH(A2,Altas!A:A,0))',
    '=INDEX(Órdenes!M:M,MATCH(B2,Órdenes!C:C,0))',
    '=INDEX(Órdenes!N:N,MATCH(B199,Órdenes!C:C,0))',
    '=D2*F2',
    '=D2/E2',
    '=INDEX(Órdenes!E:E,MATCH(B2,Órdenes!C:C,0))',
    '=INDEX("Información fija"!$C$3:$C$7,MATCH(I2,"Información fija"!$B$3:$B$7,0))',
    '="Cualquier reacción adversa que se presente con el uso de este medicamento, agradeceremos sea notificado de forma inmediata a Eseotres Pharma a farmacovigilancia@eseotrespharma.com, página web eseotrespharma.com o vía telefónica al 7356882634"&CHAR(10)&"Número de proveedor: 0000150462"&CHAR(10)&"Número de enlace SAI: "&C2&CHAR(10)&"Número de contrato: "&INDEX("Información fija"!$B$9:$B$12,MATCH(C2,"Información fija"!$C$9:$C$12,0))&CHAR(10)&"Clave: "&I2&CHAR(10)&"Denominación Social de la Afianzadora: Sofimex, Institución de Garantías, S.A"&CHAR(10)&"Número de fianza: "&INDEX("Información fija"!$D$9:$D$12,MATCH(C2,"Información fija"!$C$9:$C$12,0))&CHAR(10)&"Número de alta: "&A2&CHAR(10)&"Número de orden de reposición: "&B2&CHAR(10)&"Lugar de entrega: "&INDEX(Órdenes!H:H,MATCH(B2,Órdenes!C:C,0))&CHAR(10)&CHAR(10)&"El PROVEEDOR cuenta con opiniones positivas y vigentes en materia de aportaciones de Seguridad Social ante El INSTITUTO e INFONAVIT, así como de obligaciones fiscales ante el SAT"',
    '=IFNA(INDEX("Reporte Paq"!B:B,MATCH(A2,"Reporte Paq"!D:D,0)),"Por facturar")'
]

for row in range(2, last_row+1):  # Starting from the second row (skip headers)
    for col, formula in enumerate(formulas, start=2):  # Starting from the second column (column B)
        cell = tigre_excel_de_trabajo.cell(row=row, column=col)
        cell.value = formula.replace("2", str(row))  # Adjust the formula to refer to the current row
"""

# Save the tigre workbook
tigre_wb.save("tigre_2024.xlsx")
