import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re

def clear_or_recreate_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    wb.create_sheet(sheet_name)

def clean_folio(folio):
    # Convert to string
    folio = str(folio)
    # Remove everything after a dot (.)
    folio = folio.split('.')[0]
    # Remove special characters
    folio = re.sub(r'[^\w\s]', '', folio)
    return folio

def copy_data_between_sheets(source_path, source_sheet_name, dest_path, dest_sheet_name):
    source_wb = load_workbook(source_path)
    source_sheet = source_wb[source_sheet_name]
    df = pd.DataFrame(source_sheet.values)
    headers = df.iloc[0]
    df = df[1:]
    df.columns = headers
    print("Dejando solamente facturas vigentes") 
    df = df[df['UUID Descripción'] == 'Vigente']
    df['Folio'] = df['Folio'].apply(clean_folio)

    df['Factura'] = df['Serie'] + "-" + df['Folio']
    
    dest_wb = load_workbook(dest_path)
    clear_or_recreate_sheet(dest_wb, dest_sheet_name)
    dest_sheet = dest_wb[dest_sheet_name]
    for r in dataframe_to_rows(df, index=False, header=True):
        dest_sheet.append(r)
    dest_wb.save(dest_path)

def add_formulas_and_headers(wb_path, sheet_name, headers_formulas, orders_sheet, order_column):
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    orders_ws = wb[orders_sheet]
    orders_df = pd.DataFrame(orders_ws.values)
    headers = orders_df.iloc[0]
    orders_df = orders_df[1:]
    orders_df.columns = headers

    # Extract the 'NÚMERO DE ORDEN DE SUMINISTRO' column
    order_numbers = orders_df[order_column].tolist()

    # Add headers
    ws.append([header for header, _ in headers_formulas])
    
    # Fill in the 'NÚMERO DE ORDEN DE SUMINISTRO' and apply formulas
    for row_idx, order_number in enumerate(order_numbers, start=2):
        ws.cell(row=row_idx, column=1).value = order_number
        for col_idx, (_, formula) in enumerate(headers_formulas, start=1):
            if formula:  # Check if there is a formula to apply
                corrected_formula = formula.replace("{row}", str(row_idx))  # Correct the formula
                ws.cell(row=row_idx, column=col_idx).value = '=' + corrected_formula  # Explicitly set formula with '='

    wb.save(wb_path)


# Define file paths and sheet names
source_file_path = "C:/Users/armjorge/Dropbox/FACT 2024/Consola IMSSB.xlsx"
dest_file_path = "./INSABI_ordenes-contratos-sellos-remisiones.xlsx"
source_sheet = "Reporte PAQ"
dest_sheet = "Reporte PAQ"
consola_sheet = "Consola"
ordenes_sheet = "Órdenes"

# Copy data from source to destination sheet
copy_data_between_sheets(source_file_path, source_sheet, dest_file_path, dest_sheet)
print(f"\n*******************************\n'PAQ' Actualizado en el {dest_file_path}\n*******************************\n")

# Clear or recreate the 'Consola' sheet
wb = load_workbook(dest_file_path)
clear_or_recreate_sheet(wb, consola_sheet)
wb.save(dest_file_path)

# Define headers and their formulas
headers_formulas = [
    ("NÚMERO DE ORDEN DE SUMINISTRO", None),  # This will be populated from another sheet
    ("ORDEN DE REMISIÓN", 'IFNA(INDEX(Remisiones!B:B,MATCH(A{row},Remisiones!C:C,0)),"Esperando remisiones")'),
    ("CLUES", 'INDEX(Órdenes!L:L,MATCH(A{row},Órdenes!G:G,0))'), 
    ("Estatus", 'INDEX(Órdenes!O:O,MATCH(A{row},Órdenes!G:G,0))'), 
    ("Contrato", 'INDEX(Contratos!A:A,MATCH(A{row},Contratos!B:B,0))'), 
    ("Fecha de expedición", 'TEXT(INDEX(Órdenes!H:H,MATCH(A{row},Órdenes!G:G,0)),"dd/mm/aaaa")'), 
    ("Fecha capturada remisión", 'IFNA(TEXT(INDEX(Remisiones!F:F,MATCH(A{row},Remisiones!C:C,0)),"dd/mm/aaaa"),"Esperando remisiones")'),
    ("Fecha de sello", 'IFNA(TEXT(INDEX(Sellos!B:B,MATCH(A{row},Sellos!A:A,0)),"dd/mm/aaaa"),"Esperando remisiones")'), 
    ("Receptor CFDI", 'IF(E{row}="LA-E115-2022-MED-INSABI-122-2023/2024","IMSS Bienestar",IF(OR(E{row}="LA-E115-2022-MED-INSABI-034-2023/2024",E{row}="LA-E115-2022-MED-INSABI-188-2023/2024"),"IMSS Bienestar","Contrato no identificado"))'), 
    ("RFC Receptor CFDI", '=IF(E{row}="LA-E115-2022-MED-INSABI-122-2023/2024","SSI220901JS5",IF(OR(E{row}="LA-E115-2022-MED-INSABI-034-2023/2024",E{row}="LA-E115-2022-MED-INSABI-188-2023/2024"),"SSI220901JS5","Contrato no identificado"))'),
    ("USO CFDI", 'IF(E{row}="LA-E115-2022-MED-INSABI-122-2023/2024","G01",IF(OR(E{row}="LA-E115-2022-MED-INSABI-034-2023/2024",E{row}="LA-E115-2022-MED-INSABI-188-2023/2024"),"G03","Contrato no identificado"))'), 
    ("Fianza", 'IF(E{row}="LA-E115-2022-MED-INSABI-188-2023/2024","2757897", IF(E{row}="LA-E115-2022-MED-INSABI-034-2023/2024","2757877",IF(E{row}="LA-E115-2022-MED-INSABI-122-2023/2024","2757888")))'), 
    ("Factura", 'INDEX(\'Reporte PAQ\'!L:L,MATCH(A{row},\'Reporte PAQ\'!H:H,0))'),
    ("Piezas", 'INDEX(Órdenes!N:N,MATCH(A{row},Órdenes!G:G,0))'), 
    ("Precio Unitario", 'INDEX(Precios!B:B,MATCH(LEFT(Q{row},15),Precios!A:A,0))'), 
    ("Total", 'N{row}*O{row}'), 
    ("Producto", 'INDEX(Órdenes!J:J,MATCH(A{row},Órdenes!G:G,0))&" "&INDEX(Órdenes!K:K,MATCH(A{row},Órdenes!G:G,0))'), 
    ("Piezas", 'INDEX(Órdenes!N:N,MATCH(A{row},Órdenes!G:G,0))'), 
    ("Lote", 'INDEX(Remisiones!I:I,MATCH(A{row},Remisiones!C:C,0))'), 
    ("Caducidad", 'INDEX(Remisiones!J:J,MATCH(A{row},Remisiones!C:C,0))'), 
    ("Leyenda PDF", '"NOS: "&A{row}&CHAR(10)&"OR: "&B{row}&CHAR(10)&"CLUES: "&C{row}&CHAR(10)&"CONTRATO: "&E{row}&CHAR(10)&"FEEX: "&F{row}&CHAR(10)&"FEEN: "&G{row}&CHAR(10)&"FERE: "&H{row}&CHAR(10)&CHAR(10)&"FIANZA: "&L{row}&CHAR(10)&"Convenio Modificatorio: PRIMER CONVENIO MODIFICATORIO DEL CONTRATO "&E{row}&CHAR(10)&"Nombre de la afianzadora: SOFIMEX, INSTITUCIÓN DE GARANTÍAS, S.A."&CHAR(10)&CHAR(10)&IF(E{row}="LA-E115-2022-MED-INSABI-122-2023/2024","DATOS BANCARIOS"&CHAR(10)&"Banco: BBVA MEXICO, S.A."&CHAR(10)&"Nombre del Beneficiario (Razón social): Eseotres Pharma, S. A. P. I de C. V."&CHAR(10)&"Número de cuenta: 0110260568"&CHAR(10)&"Número de CLABE: 012180001102605689"&CHAR(10)&"Número de sucursal: 0843"&CHAR(10)&"SWIFT CODE: BCMRMXMMXXX"&CHAR(10)&"Factura con cargo al Fondo de Salud para el Bienestar","")'), 
    ("Addenda XML", '"<cfdi:Addenda>"&CHAR(10)&"<REM>"&CHAR(10)&"<NOS>"&A{row}&"</NOS>"&CHAR(10)&"<OR>"&B{row}&"</OR>"&CHAR(10)&"<CLUES>"&C{row}&"</CLUES>"&CHAR(10)&"<CONTRATO>"&E{row}&"</CONTRATO>"&CHAR(10)&"<FEEX>"&F{row}&"</FEEX>"&CHAR(10)&"<FEEN>"&G{row}&"</FEEN>"&CHAR(10)&"<FERE>"&H{row}&"</FERE>"&CHAR(10)&"</REM>"&CHAR(10)&"</cfdi:Addenda>"'),
    ("Cantidad pagada",None)
]
print(f"\n*******************************\n Asegúrate de revisar las fórmulas de la consola {dest_file_path}\n*******************************\n")

# Add headers and 'NÚMERO DE ORDEN DE SUMINISTRO' from 'Órdenes' and other formulas to the 'Consola' sheet
add_formulas_and_headers(dest_file_path, consola_sheet, headers_formulas, ordenes_sheet, 'NÚMERO DE ORDEN DE SUMINISTRO')
